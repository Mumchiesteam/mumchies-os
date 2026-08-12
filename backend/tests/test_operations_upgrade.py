from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import pytest
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.api.routes.couriers import PackageDetailsPayload, _build_delhivery_payload, _build_shiprocket_order_payload
from app.api.routes.analytics import _payment
from app.api.routes import orders as orders_routes
from app.api.routes.labels import label_queue
from app.api.routes.orders import AddressValidationPayload, ExportPayload, export_orders, validate_order_address
from app.db.base import Base
from app.models.shiprocket import ShiprocketShipment
from app.repositories.shiprocket import upsert_shipment
from app.services import label_printing, order_operations
from app.services.label_printing import LabelPrintError, confirm_batch, create_batch, print_ready_pdf
from app.services.order_operations import OrderOperationsStore
from app.services.shopify import ShopifyService


@pytest.fixture()
def db(tmp_path: Path):
    engine = create_engine(f"sqlite+pysqlite:///{tmp_path / 'ops.db'}")
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine, autoflush=False, autocommit=False)()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


def raw_order(status="partially_paid", outstanding="1378.00"):
    return {
        "id": 1, "name": "#321607", "order_number": 321607, "created_at": "2026-07-20T05:22:00Z",
        "customer": {}, "shipping_address": {"name": "Customer", "address1": "12 Road", "city": "Delhi", "province": "Delhi", "zip": "110001"},
        "line_items": [{"title": "Product", "sku": "SKU", "quantity": 1, "grams": 950, "price": "1577.00"}],
        "shipping_lines": [], "total_price": "1577.00", "current_total_price": "1577.00",
        "total_outstanding": outstanding, "financial_status": status, "fulfillment_status": None,
        "cancelled_at": None, "tags": "", "payment_gateway_names": ["COD", "PayU"],
        "_transaction_summary": [{"kind": "sale", "status": "success", "amount": "199.00", "gateway": "PayU"}],
    }


def pdf(width=288, height=432, pages=1):
    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=width, height=height)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def test_partial_payment_normalization_uses_outstanding_balance():
    order = ShopifyService._to_order(raw_order())
    assert (order.order_total, order.paid_amount, order.cod_collectable_amount, order.payment_type) == (1577, 199, 1378, "partial_cod")


def test_shopify_shipping_address_preserves_all_drawer_fields():
    value = raw_order()
    value["shipping_address"] = {
        "name": "Test Customer", "phone": "+919999999999",
        "address1": "12 Main Road", "address2": "Second Floor",
        "city": "Delhi", "province": "Delhi", "zip": "110001",
    }
    address = ShopifyService._to_order(value).shipping_address
    assert address.model_dump() == {
        "name": "Test Customer", "phone": "+919999999999",
        "address": "12 Main Road Second Floor",
        "address_line1": "12 Main Road", "address_line2": "Second Floor",
        "landmark": None, "city": "Delhi", "state": "Delhi", "pincode": "110001",
    }


@pytest.mark.parametrize(
    ("status", "outstanding", "gateways", "payment_type"),
    [("pending", "1577", ["Cash on Delivery (COD)"], "cod"), ("paid", "0", ["PayU"], "prepaid")],
)
def test_full_cod_and_prepaid_regression(status, outstanding, gateways, payment_type):
    value = raw_order(status, outstanding)
    value["payment_gateway_names"] = gateways
    assert ShopifyService._to_order(value).payment_type == payment_type


def test_cancelled_zero_balance_cod_keeps_stable_payment_type():
    value = raw_order("voided", "0")
    value["payment_gateway_names"] = ["Cash on Delivery (COD)"]
    value["cancelled_at"] = "2026-07-21T05:22:00Z"
    assert ShopifyService._to_order(value).payment_type == "cod"


def test_cancelled_payu_remains_prepaid():
    value = raw_order("voided", "0")
    value["payment_gateway_names"] = ["PayU"]
    value["cancelled_at"] = "2026-07-21T05:22:00Z"
    assert ShopifyService._to_order(value).payment_type == "prepaid"


def test_cancelled_partial_cod_uses_retained_gateway_and_payment_evidence():
    value = raw_order("voided", "0")
    value["cancelled_at"] = "2026-07-21T05:22:00Z"
    assert ShopifyService._to_order(value).payment_type == "partial_cod"


def test_cancelled_zero_balance_cod_counts_as_cod_not_prepaid_cancellation():
    active_cod = raw_order("pending", "1577")
    active_cod["payment_gateway_names"] = ["Cash on Delivery (COD)"]
    cancelled_cod = raw_order("voided", "0")
    cancelled_cod["payment_gateway_names"] = ["Cash on Delivery (COD)"]
    cancelled_cod["cancelled_at"] = "2026-07-21T05:22:00Z"
    active_payu = raw_order("paid", "0")
    active_payu["payment_gateway_names"] = ["PayU"]
    payment = {row["key"]: row for row in _payment([
        ShopifyService._to_order(active_cod),
        ShopifyService._to_order(cancelled_cod),
        ShopifyService._to_order(active_payu),
    ])}
    assert payment["cod"]["orders"] == 1
    assert payment["cod"]["cancellation_percent"] == 50
    assert payment["prepaid"]["cancellation_percent"] == 0


def test_partial_cod_provider_payloads_collect_only_balance():
    order = ShopifyService._to_order(raw_order())
    operations = {"verified_address_snapshot": {"customer_name": "Customer", "phone": "9999999999", "address_line1": "12 Road", "address_line2": "", "landmark": "", "city": "Delhi", "state": "Delhi", "pincode": "110001"}}
    package = PackageDetailsPayload(weight_kg=.95, length_cm=5, breadth_cm=5, height_cm=5)
    assert _build_delhivery_payload(order, operations, package)["cod_amount"] == 1378
    assert _build_shiprocket_order_payload(order, operations, package)["sub_total"] == 1378


def test_first_human_action_is_persistent_and_automated_sync_does_not_count(tmp_path, monkeypatch):
    path = tmp_path / "operations.json"
    monkeypatch.setattr(order_operations, "OPS_FILE", path)
    OrderOperationsStore.save_address_sync_results("1", {"shopify_order": "synced"})
    assert OrderOperationsStore.get("1").get("first_action_at") is None
    OrderOperationsStore.append_call_log("1", {"result": "No Answer", "timestamp": "2026-07-21T10:00:00", "operator": "A", "comment": ""})
    assert OrderOperationsStore.get("1")["first_action_at"] == "2026-07-21T10:00:00"


@pytest.mark.anyio
async def test_address_hard_blockers_and_advisory_warnings(db):
    missing_pin = await validate_order_address("1", AddressValidationPayload(address_line1="12 Main Road"), db)
    assert missing_pin["valid"] is False
    warning = await validate_order_address("1", AddressValidationPayload(address_line1="Main Road Area", city="Delhi", state="Delhi", pincode="110001"), db)
    assert warning["valid"] is True
    assert "House or flat number was not detected" in warning["warnings"]
    assert "Landmark is missing" in warning["warnings"]
    assert warning["shiprocket_confidence_score"] is None


def test_print_ready_pdf_is_exact_4x6_and_preserves_page_count():
    result = print_ready_pdf(pdf(pages=2))
    pages = PdfReader(BytesIO(result)).pages
    assert len(pages) == 2
    assert all((float(page.mediabox.width), float(page.mediabox.height)) == (288, 432) for page in pages)


def test_a4_without_provider_crop_box_is_rejected():
    with pytest.raises(LabelPrintError, match="no provider-defined 4 x 6"):
        print_ready_pdf(pdf(595, 842))


@pytest.mark.anyio
async def test_batch_generation_does_not_mark_printed_and_prevents_duplicate(db, monkeypatch, tmp_path):
    monkeypatch.setattr(label_printing, "LABEL_DIR", tmp_path)
    async def fake_label(_shipment):
        return pdf()
    monkeypatch.setattr(label_printing, "official_label", fake_label)
    upsert_shipment(db, "1", provider="shiprocket", awb="A1", shipment_id="1", booking_status="booked", label_print_status="not_printed")
    batch = await create_batch(db, ["1"], "Operator")
    assert db.get(ShiprocketShipment, "1").label_print_status == "awaiting_confirmation"
    with pytest.raises(LabelPrintError, match="active print batch"):
        await create_batch(db, ["1"], "Operator")
    confirm_batch(db, batch.id, {"1"}, "Operator")
    assert db.get(ShiprocketShipment, "1").label_print_count == 1
    confirm_batch(db, batch.id, {"1"}, "Operator")
    assert db.get(ShiprocketShipment, "1").label_print_count == 1


@pytest.mark.anyio
async def test_partial_confirmation_returns_failed_item_to_queue(db, monkeypatch, tmp_path):
    monkeypatch.setattr(label_printing, "LABEL_DIR", tmp_path)
    async def fake_label(_shipment):
        return pdf()
    monkeypatch.setattr(label_printing, "official_label", fake_label)
    for identifier in ("1", "2"):
        upsert_shipment(db, identifier, provider="shiprocket", awb=f"A{identifier}", shipment_id=identifier, booking_status="booked", label_print_status="not_printed")
    batch = await create_batch(db, ["1", "2"], "Operator")
    confirm_batch(db, batch.id, {"1"}, "Operator")
    assert db.get(ShiprocketShipment, "1").label_print_status == "printed"
    assert db.get(ShiprocketShipment, "2").label_print_status == "not_printed"


def test_legacy_booked_shipment_remains_untracked(db):
    shipment = ShiprocketShipment(order_id="legacy", provider="delhivery", awb="OLD", booking_status="booked")
    db.add(shipment)
    db.commit()
    assert shipment.label_print_status is None


@pytest.mark.anyio
async def test_print_history_does_not_remove_booked_shipments_from_ready_to_ship(db):
    from datetime import datetime, timedelta, timezone

    now = datetime.now(timezone.utc)
    db.add_all([
        ShiprocketShipment(order_id="today", provider="delhivery", awb="T", booking_status="booked", label_print_status="printed", label_last_printed_at=now, label_last_printed_by="Operator"),
        ShiprocketShipment(order_id="old", provider="delhivery", awb="O", booking_status="booked", label_print_status="printed", label_last_printed_at=now - timedelta(days=2), label_last_printed_by="Operator"),
    ])
    db.commit()
    result = await label_queue(db)
    assert {item["order_id"] for item in result["ready_to_ship"]} == {"today", "old"}
    assert result["manifested"] == []


async def response_bytes(response) -> bytes:
    return b"".join([chunk async for chunk in response.body_iterator])


@pytest.mark.anyio
async def test_current_view_export_respects_selected_ids(db, monkeypatch):
    first = ShopifyService._to_order(raw_order()).model_copy(update={"order_id": "1"})
    second = ShopifyService._to_order({**raw_order("paid", "0"), "id": 2, "name": "#2"}).model_copy(update={"order_id": "2", "order_number": "2"})
    async def fake_orders(_db):
        return [first, second]
    monkeypatch.setattr(orders_routes, "_load_orders", fake_orders)
    workbook = load_workbook(BytesIO(await response_bytes(await export_orders(ExportPayload(mode="current", order_ids=["1"]), db))))
    assert workbook["Current View"].max_row == 2
    assert workbook["Current View"]["A2"].value == "321607"


@pytest.mark.anyio
async def test_full_export_has_required_tabs_and_payment_columns(db, monkeypatch):
    order = ShopifyService._to_order(raw_order()).model_copy(update={"order_id": "1"})
    async def fake_orders(_db):
        return [order]
    monkeypatch.setattr(orders_routes, "_load_orders", fake_orders)
    workbook = load_workbook(BytesIO(await response_bytes(await export_orders(ExportPayload(mode="full"), db))))
    required = {"Summary", "All Orders", "Fresh Orders", "Previous Pending", "COD", "Partial COD", "Prepaid", "High Risk", "Repeat Customers"}
    assert required.issubset(workbook.sheetnames)
    headings = [cell.value for cell in workbook["All Orders"][1]]
    assert {"Total Value", "Amount Paid", "COD / Outstanding", "Payment Type"}.issubset(headings)


@pytest.mark.anyio
async def test_full_export_fresh_tab_uses_canonical_fresh_classification(db, monkeypatch):
    fresh = ShopifyService._to_order({**raw_order(), "created_at": datetime.now(timezone.utc).isoformat()}).model_copy(update={"order_id": "1", "order_number": "1"})
    cancelled = ShopifyService._to_order({**raw_order(), "id": 2, "name": "#2", "cancelled_at": "2026-08-08T10:00:00Z"}).model_copy(update={"order_id": "2", "order_number": "2"})
    async def fake_orders(_db): return [fresh, cancelled]
    monkeypatch.setattr(orders_routes, "_load_orders", fake_orders)
    workbook = load_workbook(BytesIO(await response_bytes(await export_orders(ExportPayload(mode="full"), db))))
    assert workbook["Fresh Orders"].max_row == 2
    assert workbook["Fresh Orders"]["A2"].value == "1"
    assert workbook["Summary"]["B3"].value == 1
