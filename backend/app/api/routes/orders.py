import asyncio
import logging
import time
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import re
from zoneinfo import ZoneInfo

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import select

from app.db.session import SessionLocal, get_db
from app.core.identity import current_actor
from app.schemas.orders import ShopifyOrder
from app.repositories.shiprocket import get_shipment, get_shipments_by_order_id, snapshot as shipment_snapshot, sync_engage_orders
from app.services.order_operations import OrderOperationsStore
from app.services.report_snapshots import ReportSnapshotStore
from app.services.delhivery import DelhiveryError, DelhiveryService
from app.services.shipment_status import derive_operational_status, has_existing_shipment_evidence, has_persisted_provider_booking_evidence, merge_shopify_fulfillment_evidence
from app.services.shiprocket import ShiprocketAPIError, ShiprocketConfigurationError, ShiprocketService
from app.services.shopify import ShopifyConfigurationError, ShopifyService, ShopifySyncError
from app.services.shopify_fulfillment import ShopifyFulfillmentSynchronizer, ShopifyFulfillmentSyncError

router = APIRouter(prefix="/orders", tags=["orders"])
RECONCILIATION_SNAPSHOT_KEY = "reconciliation"
_reconciliation_refresh_task: asyncio.Task[None] | None = None


class AddressPayload(BaseModel):
    operator: str | None = None
    customer_name: str | None = None
    phone: str | None = None
    address_line1: str | None = None
    address_line2: str | None = None
    landmark: str | None = None
    city: str | None = None
    state: str | None = None
    pincode: str | None = None
    courier_sync_status: str | None = None
    courier_sync_error: str | None = None
    verified_by: str | None = None
    update_customer_address: bool = True
    one_time_delivery_address: bool = False
    use_as_default_address: bool = False


class CallLogPayload(BaseModel):
    result: str = Field(...)
    timestamp: str | None = None
    operator: str | None = None
    comment: str | None = None


class AddressConfirmationPayload(BaseModel):
    comment: str = ""
    operator: str | None = None


class SaveVerifyAddressPayload(AddressPayload):
    operator: str | None = None


class CancellationPayload(BaseModel):
    operator: str | None = None
    comment: str | None = None
    cancel_shopify: bool = True
    cancel_shiprocket: bool = True


class ShiprocketOnlyCancellationPayload(BaseModel):
    shiprocket_order_id: str
    order_number: str
    operator: str | None = None


class VerifyAddressPayload(BaseModel):
    operator: str | None = None
    verified_at: str | None = None
    address_snapshot: dict[str, str | None] = Field(default_factory=dict)


class ExportPayload(BaseModel):
    mode: str = "current"
    order_ids: list[str] = Field(default_factory=list)


class AddressValidationPayload(BaseModel):
    address_line1: str | None = None
    address_line2: str | None = None
    landmark: str | None = None
    city: str | None = None
    state: str | None = None
    pincode: str | None = None


class OrdersPage(BaseModel):
    items: list[ShopifyOrder]
    page: int
    page_size: int
    total: int
    total_pages: int
    counts: dict[str, int | float]


def _merged_operational_state(order: ShopifyOrder, operations: dict[str, object]) -> ShopifyOrder:
    call_logs = operations.get("call_logs") or []
    human_actions = operations.get("human_actions") or []
    first_action_at = operations.get("first_action_at")
    if not first_action_at and call_logs:
        first_action_at = min((str(value.get("timestamp")) for value in call_logs if value.get("timestamp")), default=None)
    if not first_action_at and any((operations.get("corrected_address"), operations.get("address_verified"), operations.get("package_details"), operations.get("selected_courier"))):
        first_action_at = "historic"

    # Single authoritative precedence chain - see app/services/shipment_status.py. Shipment/
    # fulfilment-backed states (Cancelled, Delivered, Shipped, Booked, NDR) always outrank
    # locally-derived operational states, so call logs/address edits can never downgrade them.
    operational_status = derive_operational_status(order, operations, operations.get("shipment"))
    latest_call = call_logs[0]["result"] if call_logs else None

    return order.model_copy(update={
        "latest_call_result": latest_call,
        "operational_status": operational_status,
        "address_verified": bool(operations.get("address_verified")),
        "address_verified_at": operations.get("address_verified_at"),
        "address_verified_by": operations.get("address_verified_by"),
        "verified_address_snapshot": operations.get("verified_address_snapshot"),
        "corrected_address": operations.get("corrected_address"),
        "courier_sync_status": operations.get("courier_sync_status"),
        "courier_sync_error": operations.get("courier_sync_error"),
        "address_sync_results": operations.get("address_sync_results"),
        "package_details": operations.get("package_details"),
        "selected_courier": operations.get("selected_courier"),
        "shipment": operations.get("shipment"),
        "first_action_at": first_action_at,
        "human_action_count": len(human_actions) or (1 if first_action_at else 0),
        "call_attempt_count": len(call_logs),
        "engage_order_id": (operations.get("shipment") or {}).get("engage_order_id"),
        "order_confirmation": (operations.get("shipment") or {}).get("order_confirmation"),
        "order_confirmation_message": (operations.get("shipment") or {}).get("order_confirmation_message"),
        "address_confirmation": (operations.get("shipment") or {}).get("address_confirmation"),
        "address_confirmation_message": (operations.get("shipment") or {}).get("address_confirmation_message"),
        "cod_to_prepaid": (operations.get("shipment") or {}).get("cod_to_prepaid"),
        "cod_to_prepaid_message": (operations.get("shipment") or {}).get("cod_to_prepaid_message"),
        "engage_last_synced_at": (operations.get("shipment") or {}).get("engage_last_synced_at"),
    })


async def _load_orders(db: Session, *, force_refresh: bool = False) -> list[ShopifyOrder]:
    try:
        orders = await ShopifyService().get_latest_orders(force_refresh=force_refresh)
        operations_map = OrderOperationsStore.all()
        shipments = get_shipments_by_order_id(db, [order.order_id for order in orders])
        merged_orders: list[ShopifyOrder] = []
        for order in orders:
            shipment = shipments.get(order.order_id)
            local_snapshot = shipment_snapshot(shipment) if shipment else None
            operations = {**operations_map.get(order.order_id, {}), "shipment": merge_shopify_fulfillment_evidence(local_snapshot, order.external_tracking)}
            merged_orders.append(_merged_operational_state(order, operations))
        return merged_orders
    except ShopifyConfigurationError as error:
        raise HTTPException(status_code=503, detail=str(error)) from error
    except httpx.HTTPStatusError as error:
        raise HTTPException(status_code=502, detail="Shopify could not provide orders. Check the store, token, and API version.") from error
    except httpx.HTTPError as error:
        raise HTTPException(status_code=502, detail="Unable to reach Shopify.") from error


async def _load_reconciliation_orders(db: Session) -> list[ShopifyOrder]:
    """Load active unfulfilled Shopify orders independently of the Orders queue lookback."""
    try:
        orders = await ShopifyService().get_active_unfulfilled_orders()
        operations_map = OrderOperationsStore.all()
        shipments = get_shipments_by_order_id(db)
        merged_orders: list[ShopifyOrder] = []
        for order in orders:
            shipment = shipments.get(order.order_id)
            local_snapshot = shipment_snapshot(shipment) if shipment else None
            operations = {**operations_map.get(order.order_id, {}), "shipment": merge_shopify_fulfillment_evidence(local_snapshot, order.external_tracking)}
            merged_orders.append(_merged_operational_state(order, operations))
        return merged_orders
    except ShopifyConfigurationError as error:
        raise HTTPException(status_code=503, detail=str(error)) from error
    except httpx.HTTPStatusError as error:
        raise HTTPException(status_code=502, detail="Shopify could not provide unfulfilled orders.") from error
    except httpx.HTTPError as error:
        raise HTTPException(status_code=502, detail="Unable to reach Shopify.") from error


async def _canonical_shipment_readback(order: ShopifyOrder, operations: dict[str, object], db: Session) -> dict[str, object] | None:
    """Read through to Shiprocket only to repair a confirmed, correlated shipment.

    This path never creates an order, assigns a courier, or requests an AWB.  It may only read
    an upstream order with the exact Shopify order number and persist identifiers already
    returned by that provider.
    """
    local = get_shipment(db, order.order_id)
    local_snapshot = shipment_snapshot(local) if local else None
    merged = merge_shopify_fulfillment_evidence(local_snapshot, order.external_tracking)
    selected = operations.get("selected_courier") if isinstance(operations.get("selected_courier"), dict) else {}
    provider = str((merged or {}).get("provider") or getattr(order.external_tracking, "provider", None) or selected.get("provider") or "").casefold()
    fulfillment = str(order.fulfillment_status or "").casefold()
    confirmed_context = bool(
        has_persisted_provider_booking_evidence(local_snapshot)
        or getattr(order.external_tracking, "awb", None)
        or fulfillment in {"fulfilled", "shipped", "partial", "partially_fulfilled"}
    )
    required = ("provider", "courier_name", "awb", "shipment_id", "provider_order_id", "booking_status", "booked_at", "latest_status")
    incomplete = not merged or any(not (merged or {}).get(field) for field in required)
    if provider == "shiprocket" and confirmed_context and incomplete:
        try:
            reconciled = await ShiprocketService().reconcile_existing_shipment(
                db, order.order_id, order.order_number, local.shipment_id if local else None,
            )
            merged = merge_shopify_fulfillment_evidence(reconciled, order.external_tracking)
            if merged is not None:
                merged["readback_reconciliation_status"] = "reconciled"
        except (ShiprocketAPIError, ShiprocketConfigurationError, httpx.HTTPError) as error:
            if merged is not None:
                merged["readback_reconciliation_status"] = "unavailable"
                merged["readback_reconciliation_error"] = str(error)
    return merged


def _requires_reconciliation_action(order: ShopifyOrder) -> bool:
    fulfillment_status = str(order.fulfillment_status or "unfulfilled").casefold()
    if fulfillment_status != "unfulfilled" or order.cancelled_at or str(order.shopify_status or "").casefold() in {"cancelled", "canceled"}:
        return False
    shipment = order.shipment or None
    return not has_existing_shipment_evidence(order, {}, shipment)


def _is_inactive(order: ShopifyOrder) -> bool:
    text = " ".join(filter(None, [
        order.fulfillment_status, order.shopify_status, order.cancelled_at,
        " ".join(order.tags), str(order.external_tracking.status if order.external_tracking else ""),
    ])).casefold()
    return bool(order.cancelled_at) or str(order.operational_status or "").casefold() in {"cancelled", "shipped", "delivered"} or any(value in text for value in ("cancel", "fulfilled", "shipped", "picked up", "dispatched", "in transit", "out for delivery", "delivered"))


def _has_pending_exception(order: ShopifyOrder) -> bool:
    shipment = order.shipment or {}
    booking_status = str(shipment.get("booking_status") or "").casefold()
    latest_status = str(shipment.get("latest_status") or "").casefold()
    failure_words = ("fail", "error", "exception", "uncertain", "unknown", "pending", "partial")
    booking_problem = booking_status in {"pending_awb", "awb_failed", "manifested_pending_waybill", "manifest_unknown", "manifest_partial"} or any(word in latest_status for word in failure_words)
    fulfillment_problem = str(shipment.get("shopify_fulfillment_sync_status") or "").casefold() == "failed" or bool(shipment.get("shopify_fulfillment_sync_error"))
    address_problem = str(shipment.get("address_sync_status") or "").casefold() == "failed" or bool(shipment.get("address_sync_error"))
    local_sync_problem = bool(order.courier_sync_error) or str(order.courier_sync_status or "").casefold() == "failed"
    address_results = order.address_sync_results or {}
    local_address_problem = any(str(value).casefold() == "failed" for key, value in address_results.items() if key != "errors")
    return booking_problem or fulfillment_problem or address_problem or local_sync_problem or local_address_problem


def _requires_operational_action(order: ShopifyOrder) -> bool:
    if _is_inactive(order) or str(order.latest_call_result or "").casefold() == "cancelled":
        return False
    shipment = order.shipment or {}
    successfully_booked = str(shipment.get("booking_status") or "").casefold() in {"booked", "complete", "completed", "awb_assigned"} and bool(shipment.get("awb"))
    if successfully_booked and not _has_pending_exception(order):
        return False
    return True


def _order_local_date(order: ShopifyOrder) -> date | None:
    try:
        created = datetime.fromisoformat(str(order.created_date).replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        return created.astimezone(ZoneInfo("Asia/Kolkata")).date()
    except (TypeError, ValueError):
        return None


def _is_fresh_order(order: ShopifyOrder, now: datetime | None = None) -> bool:
    if not _requires_operational_action(order):
        return False
    reference = now or datetime.now(timezone.utc)
    return _order_local_date(order) == reference.astimezone(ZoneInfo("Asia/Kolkata")).date()


def _call_outcome_requires_follow_up(order: ShopifyOrder) -> bool:
    """Queue routing uses only persisted call outcomes supported by the Orders UI."""
    return (
        order.payment_type in {"cod", "partial_cod"}
        and order.call_attempt_count >= 1
        and str(order.latest_call_result or "").strip() in {
            "No Answer", "Busy", "Switched Off", "Callback Requested",
        }
    )


def _call_outcome_is_on_hold(order: ShopifyOrder) -> bool:
    return order.payment_type in {"cod", "partial_cod"} and str(order.latest_call_result or "").strip() == "On Hold"


def _engage_category(value: object) -> str:
    raw = str(value) if value is not None else ""
    return {
        "0": "pending", "1": "pending", "2": "successful", "21": "successful",
        "3": "cancelled", "6": "disabled", "NA": "disabled",
    }.get(raw, "unknown")


def _base_filtered_orders(orders: list[ShopifyOrder], search: str, payment: str, risk: str, order_confirmation: str = "all", address_verification: str = "all", cod_to_prepaid: str = "all") -> list[ShopifyOrder]:
    needle = search.strip().casefold()
    result = []
    for order in orders:
        searchable = " ".join(filter(None, [
            order.order_number, order.shopify_name, order.customer_name, order.phone,
            " ".join(f"{product.product_name} {product.sku or ''}" for product in order.products),
        ])).casefold()
        if needle and needle not in searchable:
            continue
        display_payment = {"cod": "cod", "partial_cod": "partial cod", "prepaid": "prepaid"}.get(order.payment_type, order.payment_type)
        if payment != "all" and display_payment != payment:
            continue
        tag_text = " ".join(order.tags).casefold()
        order_risk = "high" if "high" in tag_text else "medium" if "medium" in tag_text else "low"
        if risk != "all" and order_risk != risk:
            continue
        if order_confirmation != "all" and _engage_category(order.order_confirmation) != order_confirmation:
            continue
        if address_verification != "all" and _engage_category(order.address_confirmation) != address_verification:
            continue
        if cod_to_prepaid != "all" and _engage_category(order.cod_to_prepaid) != cod_to_prepaid:
            continue
        result.append(order)
    return result


def _full_counts(orders: list[ShopifyOrder], now: datetime, db: Session) -> dict[str, int | float]:
    fresh = [order for order in orders if _matches_queue(order, "fresh", now)]
    previous = [order for order in orders if _requires_operational_action(order) and not _is_fresh_order(order, now)]
    follow_up = [order for order in previous if not _call_outcome_is_on_hold(order)]
    on_hold = [order for order in previous if _call_outcome_is_on_hold(order)]
    cod = [order for order in orders if order.payment_type in {"cod", "partial_cod"}]
    prepaid = [order for order in orders if order.payment_type == "prepaid"]
    high_risk = [order for order in orders if "high" in " ".join(order.tags).casefold() and not _is_inactive(order)]
    customer_counts: dict[str, int] = {}
    for order in orders:
        if order.customer_id:
            customer_counts[order.customer_id] = customer_counts.get(order.customer_id, 0) + 1
    repeat = [order for order in orders if (order.customer_orders_count or 0) > 1 or bool(order.customer_id and customer_counts.get(order.customer_id, 0) > 1)]
    from app.models.shiprocket import ShiprocketShipment
    local_today = now.astimezone(ZoneInfo("Asia/Kolkata")).date()
    shipments = db.scalars(select(ShiprocketShipment).where(ShiprocketShipment.label_print_status.in_(["not_printed", "awaiting_confirmation", "printed"]))).all() if hasattr(db, "scalars") else []
    printed_today = [value for value in shipments if value.label_print_status == "printed" and value.label_last_printed_at and (value.label_last_printed_at if value.label_last_printed_at.tzinfo else value.label_last_printed_at.replace(tzinfo=timezone.utc)).astimezone(ZoneInfo("Asia/Kolkata")).date() == local_today]
    return {
        "operations": len(fresh) + len(previous), "fresh": len(fresh), "previous": len(previous),
        "follow_up": len(follow_up), "on_hold": len(on_hold), "all": len(orders),
        "labels_to_print": sum(1 for value in shipments if value.label_print_status == "not_printed" and value.booking_status == "booked" and value.awb),
        "awaiting_confirmation": sum(1 for value in shipments if value.label_print_status == "awaiting_confirmation"),
        "printed_today": len(printed_today), "new_orders": len(fresh),
        "cod": len(cod), "prepaid": len(prepaid), "high_risk": len(high_risk), "repeat_customers": len(repeat),
        "cod_collectable": sum(float(order.cod_collectable_amount) for order in cod),
        "prepaid_value": sum(float(order.order_total) for order in prepaid),
        "awaiting_order_confirmation": sum(1 for order in orders if _engage_category(order.order_confirmation) == "pending"),
        "awaiting_address_verification": sum(1 for order in orders if _engage_category(order.address_confirmation) == "pending"),
        "cod_conversion_pending": sum(1 for order in orders if _engage_category(order.cod_to_prepaid) == "pending"),
    }


def _matches_queue(order: ShopifyOrder, queue: str, now: datetime, pending_view: str = "follow_up") -> bool:
    shipment = order.shipment or {}
    if queue == "fresh":
        return _is_fresh_order(order, now)
    if queue == "previous":
        if not _requires_operational_action(order) or _is_fresh_order(order, now):
            return False
        return _call_outcome_is_on_hold(order) if pending_view == "on_hold" else not _call_outcome_is_on_hold(order)
    if queue == "labels_to_print":
        return shipment.get("booking_status") == "booked" and bool(shipment.get("awb")) and shipment.get("label_print_status") == "not_printed"
    if queue == "awaiting_confirmation":
        return shipment.get("label_print_status") == "awaiting_confirmation"
    if queue == "printed_today":
        printed = shipment.get("label_last_printed_at")
        return shipment.get("label_print_status") == "printed" and bool(printed) and datetime.fromisoformat(str(printed).replace("Z", "+00:00")).astimezone(ZoneInfo("Asia/Kolkata")).date() == now.astimezone(ZoneInfo("Asia/Kolkata")).date()
    return True


@router.get("", response_model=OrdersPage)
async def list_orders(
    page: int = 1,
    page_size: int = 20,
    queue: str = "all",
    search: str = "",
    payment: str = "all",
    risk: str = "all",
    sort: str = "newest",
    order_confirmation: str = "all",
    address_verification: str = "all",
    cod_to_prepaid: str = "all",
    attempt: str = "all",
    pending_view: str = "follow_up",
    db: Session = Depends(get_db),
) -> OrdersPage:
    """Return one filtered page of orders. The client never receives the unpaged collection."""
    if page < 1:
        raise HTTPException(status_code=422, detail="page must be at least 1.")
    if page_size not in {20, 50, 100}:
        raise HTTPException(status_code=422, detail="page_size must be one of 20, 50, or 100.")
    if queue not in {"fresh", "previous", "all", "labels_to_print", "awaiting_confirmation", "printed_today"}:
        raise HTTPException(status_code=422, detail="Unknown orders queue.")
    if attempt not in {"all", "1", "2", "3", "4_plus"}:
        raise HTTPException(status_code=422, detail="Unknown attempt filter.")
    if pending_view not in {"follow_up", "on_hold"}:
        raise HTTPException(status_code=422, detail="Unknown previous-pending view.")
    request_started = time.perf_counter()
    shopify_started = time.perf_counter()
    orders = await _load_orders(db)
    shopify_ms = (time.perf_counter() - shopify_started) * 1000
    now = datetime.now(timezone.utc)
    allowed_engage_filters = {"all", "pending", "successful", "cancelled", "disabled", "unknown"}
    if any(value not in allowed_engage_filters for value in (order_confirmation, address_verification, cod_to_prepaid)):
        raise HTTPException(status_code=422, detail="Unknown Engage filter.")
    base_filtered = _base_filtered_orders(orders, search, payment, risk, order_confirmation, address_verification, cod_to_prepaid)
    counts = _full_counts(base_filtered, now, db)
    effective_queue = "all" if search.strip() else queue
    filtered = [order for order in base_filtered if _matches_queue(order, effective_queue, now, pending_view)]
    if attempt != "all" and pending_view == "follow_up":
        filtered = [order for order in filtered if _call_outcome_requires_follow_up(order) and (order.call_attempt_count >= 4 if attempt == "4_plus" else order.call_attempt_count == int(attempt))]
    reverse = sort not in {"oldest", "value_asc"}
    if sort in {"value_asc", "value_desc"}:
        filtered.sort(key=lambda value: float(value.total_amount), reverse=reverse)
    elif sort in {"cod_first", "prepaid_first"}:
        preferred = "prepaid" if sort == "prepaid_first" else "cod"
        filtered.sort(key=lambda value: (value.payment_type != preferred and not (preferred == "cod" and value.payment_type == "partial_cod"), -datetime.fromisoformat(value.created_date.replace("Z", "+00:00")).timestamp()))
    else:
        filtered.sort(key=lambda value: datetime.fromisoformat(value.created_date.replace("Z", "+00:00")), reverse=reverse)
    total = len(filtered)
    total_pages = max(1, (total + page_size - 1) // page_size)
    effective_page = min(page, total_pages)
    start = (effective_page - 1) * page_size
    logging.getLogger(__name__).info(
        "orders_list total_ms=%.2f shopify_context_ms=%.2f filter_merge_ms=%.2f orders_processed=%d",
        (time.perf_counter() - request_started) * 1000, shopify_ms,
        (time.perf_counter() - request_started) * 1000 - shopify_ms, len(orders),
    )
    return OrdersPage(items=filtered[start:start + page_size], page=effective_page, page_size=page_size, total=total, total_pages=total_pages, counts=counts)


@router.get("/{order_id}/operations")
async def get_order_operations(order_id: str, db: Session = Depends(get_db)) -> dict[str, object]:
    operations = OrderOperationsStore.get(order_id)
    try:
        order = await ShopifyService().get_order(order_id)
    except (ShopifyConfigurationError, httpx.HTTPError):
        order = None
    if order is not None:
        shipment = await _canonical_shipment_readback(order, operations, db)
    else:
        local = get_shipment(db, order_id)
        shipment = shipment_snapshot(local) if local else None
    if shipment is not None:
        operations = {**operations, "shipment": shipment}
    return operations


@router.post("/{order_id}/shopify-fulfillment/sync")
async def sync_shopify_fulfillment(order_id: str, db: Session = Depends(get_db)) -> dict[str, object]:
    """Idempotently create or repair Shopify fulfillment tracking for a booked shipment."""
    try:
        result = await ShopifyFulfillmentSynchronizer().sync(
            db, order_id, f"gid://shopify/Order/{order_id}"
        )
        return {"order_id": order_id, "shipment": result}
    except ShopifyFulfillmentSyncError as error:
        raise HTTPException(status_code=409, detail=str(error)) from error


@router.post("/{order_id}/address/validate")
async def validate_order_address(order_id: str, payload: AddressValidationPayload, db: Session = Depends(get_db)) -> dict[str, object]:
    text = " ".join(filter(None, [payload.address_line1, payload.address_line2, payload.landmark])).strip()
    pincode = str(payload.pincode or "").strip()
    blockers: list[str] = []
    warnings: list[str] = []
    if not text:
        blockers.append("Address is blank")
    if not pincode:
        blockers.append("Pincode required")
    elif not re.fullmatch(r"\d{6}", pincode):
        blockers.append("Pincode must contain exactly six digits")
    if text and not re.search(r"\b(?:house|flat|plot|door|h\.?\s*no|\d+[A-Za-z/-]*)\b", text, re.IGNORECASE):
        warnings.append("House or flat number was not detected")
    if text and len(text) < 18:
        warnings.append("Address looks unusually short")
    if not payload.landmark:
        warnings.append("Landmark is missing")
    pieces = [value.casefold().strip() for value in (payload.address_line1, payload.address_line2) if value]
    if len(pieces) == 2 and pieces[0] == pieces[1]:
        warnings.append("Duplicate address text exists")
    shipment = get_shipment(db, order_id)
    # address_confidence_score/category/source (see ShiprocketShipment model) are always None
    # today. Investigated 2026-07-21: none of the Shiprocket Shipping API v1 endpoints this app
    # calls (auth/login, settings/company/pickup, courier/serviceability, orders/create/adhoc,
    # courier/assign/awb, orders search, courier/track/awb, courier/generate/label,
    # courier/awb/update - see app/services/shiprocket.py) return an address-confidence-like
    # field in their response payloads. A confidence/quality score does exist, but only under
    # Shiprocket's separate "Sense" product (Address Score / SenseAddress APIs, sense.shiprocket.in
    # per public docs) - a different, separately-licensed API with its own credentials that this
    # app does not integrate with. Do not fabricate a value here; leave these columns null until
    # Sense (or an equivalent documented endpoint) is actually integrated. The columns/API fields/
    # UI below are kept in place so a real score can be wired in later without a schema change.
    score = shipment.address_confidence_score if shipment else None
    category = shipment.address_confidence_category if shipment else None
    return {
        "valid": not blockers,
        "status": "Pincode required" if blockers else "Address has warnings" if warnings else "Address looks complete",
        "blockers": blockers,
        "warnings": warnings,
        "shiprocket_confidence_score": score,
        "shiprocket_confidence_category": category,
        "shiprocket_confidence_source": shipment.address_confidence_source if shipment else None,
        "shiprocket_message": "Shiprocket score not available" if score is None else f"Shiprocket confidence: {score:g}%" + (f" - {category}" if category else ""),
    }


def _export_row(order: ShopifyOrder) -> list[object]:
    india = ZoneInfo("Asia/Kolkata")
    created = datetime.fromisoformat(order.created_date.replace("Z", "+00:00")).astimezone(india)
    address = order.corrected_address or (order.shipping_address.model_dump() if order.shipping_address else {})
    shipment = order.shipment or {}
    return [
        order.order_number, created.date(), created.time().replace(second=0, microsecond=0), order.customer_name,
        order.phone, address.get("city"), address.get("state"), address.get("pincode"), float(order.order_total),
        float(order.paid_amount), float(order.cod_collectable_amount), order.payment_type, order.payment_status,
        "High" if "high risk" in " ".join(order.tags).casefold() else "Low", "Repeat" if (order.customer_orders_count or 0) > 1 else "New",
        order.operational_status, order.call_attempt_count, "Verified" if order.address_verified else "Pending",
        shipment.get("address_confidence_score"), shipment.get("address_confidence_category"), shipment.get("provider"),
        shipment.get("courier_name"), shipment.get("awb"), shipment.get("latest_status"), shipment.get("booked_at"),
        shipment.get("label_print_status"), shipment.get("label_last_printed_at"),
    ]


@router.post("/export")
async def export_orders(payload: ExportPayload, db: Session = Depends(get_db)):
    orders = await _load_orders(db)
    selected = orders if payload.mode == "full" else [order for order in orders if order.order_id in set(payload.order_ids)]
    headers = ["Order Number", "Order Date", "Order Time", "Customer", "Phone", "City", "State", "Pincode", "Total Value", "Amount Paid", "COD / Outstanding", "Payment Type", "Financial Status", "Risk", "Customer Type", "Operational Status", "Call Attempts", "Address Verification", "Address Confidence", "Address Category", "Courier Provider", "Courier", "AWB", "Shipment Status", "Booking Time", "Label Print Status", "Last Printed Time"]
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(name: str, values: list[ShopifyOrder]) -> None:
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for value in values:
            sheet.append(_export_row(value))
        if sheet.max_row >= 2:
            for column in (9, 10, 11):
                for cells in sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=sheet.max_row):
                    cells[0].number_format = '₹#,##0.00'
        sheet.freeze_panes = "A2"

    if payload.mode == "full":
        summary = workbook.create_sheet("Summary")
        fresh = [value for value in orders if _is_fresh_order(value)]
        summary.append(["Metric", "Count"])
        for metric, count in (("All Orders", len(orders)), ("Fresh Orders", len(fresh))):
            summary.append([metric, count])
        previous = [value for value in orders if value.first_action_at and value.operational_status not in {"Ready for Booking", "Booked", "Shipped", "Delivered", "Cancelled"}]
        tabs = {
            "All Orders": orders, "Fresh Orders": fresh, "Previous Pending": previous,
            "COD": [value for value in orders if value.payment_type in {"cod", "partial_cod"}],
            "Partial COD": [value for value in orders if value.payment_type == "partial_cod"],
            "Prepaid": [value for value in orders if value.payment_type == "prepaid"],
            "High Risk": [value for value in orders if "high risk" in " ".join(value.tags).casefold()],
            "Repeat Customers": [value for value in orders if (value.customer_orders_count or 0) > 1],
        }
        for name, values in tabs.items():
            add_sheet(name, values)
    else:
        add_sheet("Current View", selected)
    output = BytesIO()
    workbook.save(output)
    timestamp = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d-%H%M")
    return StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="mumchies-orders-{timestamp}.xlsx"'})


async def _official_shipping_label(order_id: str, db: Session, *, inline: bool = False, print_ready: bool = False):
    try:
        shipment = get_shipment(db, order_id)
        if shipment is None:
            raise HTTPException(status_code=404, detail="No courier shipment exists for this order.")
        if not shipment.awb:
            raise HTTPException(status_code=404, detail="No AWB exists for this shipment.")
        if shipment.booking_status and shipment.booking_status != "booked":
            raise HTTPException(status_code=409, detail="The shipment is not confirmed and label-eligible.")
        from app.services.label_printing import LabelPrintError, LabelService
        from app.services.courier_platform import ProviderError, courier_registry
        try:
            if print_ready:
                content, media_type, extension = await LabelService().print_ready(shipment), "application/pdf", "pdf"
            else:
                label = await courier_registry.get(str(shipment.provider or "shiprocket")).download_label(shipment_snapshot(shipment))
                content, media_type, extension = label.content, label.content_type, label.format.value
        except (LabelPrintError, ProviderError) as error:
            raise HTTPException(status_code=409, detail=str(error)) from error
        filename = f"{shipment.provider or 'courier'}-{shipment.provider_order_id or order_id}-{shipment.awb}.{extension}"
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={
                "Content-Disposition": f'{"inline" if inline else "attachment"}; filename="{filename}"',
                "Cache-Control": "private, no-store",
                "X-Content-Type-Options": "nosniff",
            },
        )
    except (DelhiveryError, ShiprocketConfigurationError, ShiprocketAPIError) as error:
        raise HTTPException(status_code=502, detail=str(error)) from error


@router.get("/{order_id}/shipment/label")
async def get_provider_shipping_label(order_id: str, disposition: str = "attachment", print_ready: bool = False, db: Session = Depends(get_db)):
    """Proxy the courier provider's official PDF bytes without re-rendering them."""
    if disposition not in {"attachment", "inline"}:
        raise HTTPException(status_code=400, detail="Label disposition must be attachment or inline.")
    return await _official_shipping_label(order_id, db, inline=disposition == "inline", print_ready=print_ready)


@router.get("/{order_id}/shipping-label")
async def get_shipping_label(order_id: str, db: Session = Depends(get_db)):
    """Backward-compatible alias for existing clients."""
    return await _official_shipping_label(order_id, db)


@router.put("/{order_id}/address")
async def update_order_address(order_id: str, payload: AddressPayload, request: Request, db: Session = Depends(get_db)) -> dict[str, object]:
    address = {
        "customer_name": payload.customer_name,
        "phone": payload.phone,
        "address_line1": payload.address_line1,
        "address_line2": payload.address_line2,
        "landmark": payload.landmark,
        "city": payload.city,
        "state": payload.state,
        "pincode": payload.pincode,
    }
    # Local correction is intentionally committed before any external write.
    OrderOperationsStore.save_address(
        order_id,
        address,
        courier_sync_status=payload.courier_sync_status,
        courier_sync_error=payload.courier_sync_error,
        operator=current_actor(request),
    )
    results: dict[str, object] = {
        "shopify_order": "failed",
        "shopify_customer": "not_applicable",
        "shiprocket": "not_applicable",
        "delhivery": "not_applicable",
        "errors": {},
    }
    service = ShopifyService()
    context: dict[str, object] | None = None
    try:
        context = await service.get_order_address_context(order_id)
    except (ShopifySyncError, httpx.HTTPError) as error:
        results["errors"]["shopify_order"] = str(error)
        if payload.update_customer_address and not payload.one_time_delivery_address:
            results["shopify_customer"] = "failed"
            results["errors"]["shopify_customer"] = "Customer address could not be resolved because the Shopify order lookup failed."
    if context is not None:
        try:
            await service.update_order_shipping_address(order_id, address)
            results["shopify_order"] = "synced"
        except (ShopifySyncError, httpx.HTTPError) as error:
            results["errors"]["shopify_order"] = str(error)

        update_customer = payload.update_customer_address and not payload.one_time_delivery_address
        customer_id = context.get("customer_id")
        if update_customer and customer_id:
            try:
                await service.update_customer_address(
                    str(customer_id),
                    context.get("shipping_address") if isinstance(context.get("shipping_address"), dict) else {},
                    address,
                    set_as_default=payload.use_as_default_address,
                )
                results["shopify_customer"] = "synced"
            except (ShopifySyncError, httpx.HTTPError) as error:
                results["shopify_customer"] = "failed"
                results["errors"]["shopify_customer"] = str(error)
        else:
            results["shopify_customer"] = "not_applicable"

    shipment = get_shipment(db, order_id)
    if shipment and shipment.awb and shipment.provider == "shiprocket":
        courier_address = {
            "shipping_customer_name": address["customer_name"],
            "shipping_phone": address["phone"],
            "shipping_address": address["address_line1"],
            "shipping_address_2": " ".join(filter(None, [address["address_line2"], address["landmark"]])),
            "shipping_city": address["city"],
            "shipping_state": address["state"],
            "shipping_pincode": address["pincode"],
        }
        try:
            await ShiprocketService().update_address(shipment.awb, courier_address)
            results["shiprocket"] = "synced"
        except (ShiprocketConfigurationError, ShiprocketAPIError, httpx.HTTPError) as error:
            results["shiprocket"] = "failed"
            results["errors"]["shiprocket"] = str(error)
    elif shipment and shipment.awb and shipment.provider == "delhivery":
        results["delhivery"] = "manual_required"
        results["errors"]["delhivery"] = "The booked Delhivery shipment was not changed automatically; cancellation/rebooking may be required."

    return OrderOperationsStore.save_address_sync_results(order_id, results)


@router.post("/{order_id}/address/save-verify")
async def save_and_verify_address(order_id: str, payload: SaveVerifyAddressPayload, request: Request, db: Session = Depends(get_db)) -> dict[str, object]:
    address = AddressValidationPayload(**payload.model_dump())
    validation = await validate_order_address(order_id, address, db)
    saved = await update_order_address(order_id, AddressPayload(**{
        **payload.model_dump(),
        "update_customer_address": True,
        "one_time_delivery_address": False,
        "use_as_default_address": False,
    }), request, db)
    verified = False
    if not validation["blockers"]:
        snapshot = {key: getattr(payload, key) for key in ("customer_name", "phone", "address_line1", "address_line2", "landmark", "city", "state", "pincode")}
        saved = OrderOperationsStore.verify_address(order_id, current_actor(request), snapshot, datetime.now(timezone.utc).isoformat())
        verified = True
    return {"operations": saved, "validation": validation, "verified": verified}


@router.post("/{order_id}/call-logs")
async def add_call_log(order_id: str, payload: CallLogPayload, request: Request) -> dict[str, object]:
    allowed = {"Confirmed", "No Answer", "Busy", "Switched Off", "On Hold", "Cancelled"}
    if payload.result not in allowed:
        raise HTTPException(status_code=422, detail="Unsupported COD result.")
    if payload.result == "On Hold" and len(str(payload.comment or "").strip()) < 3:
        raise HTTPException(status_code=422, detail="Add a reason before placing the order On Hold.")
    entry = {
        "result": payload.result,
        "timestamp": payload.timestamp or datetime.now().isoformat(timespec="seconds"),
        "operator": current_actor(request),
        "comment": payload.comment,
    }
    started = time.perf_counter()
    saved = OrderOperationsStore.append_call_log(order_id, entry)
    logging.getLogger(__name__).info("cod_call_save persistence_ms=%.2f", (time.perf_counter() - started) * 1000)
    return saved


@router.post("/{order_id}/whatsapp/cod-confirmation-opened")
async def record_cod_whatsapp_opened(order_id: str, request: Request) -> dict[str, object]:
    operations = OrderOperationsStore.get(order_id)
    logs = operations.get("call_logs") or []
    latest = logs[0].get("result") if logs and isinstance(logs[0], dict) else None
    if latest not in {"No Answer", "Busy", "Switched Off"}:
        raise HTTPException(status_code=409, detail="WhatsApp is only available after a failed COD contact result.")
    return OrderOperationsStore.record_timeline_event(
        order_id, "WhatsApp opened for COD confirmation", operator=current_actor(request)
    )


@router.post("/{order_id}/address-confirmation-comments")
async def add_address_confirmation_comment(order_id: str, payload: AddressConfirmationPayload, request: Request) -> dict[str, object]:
    return OrderOperationsStore.append_address_confirmation(order_id, payload.comment.strip(), current_actor(request), datetime.now(timezone.utc).isoformat())


async def _cancellation_preflight(order_id: str, db: Session) -> dict[str, object]:
    shipment = get_shipment(db, order_id)
    shopify = await ShopifyService().get_order_cancellation_context(order_id)
    upstream = None
    shiprocket_error = None
    shiprocket_lookup_id = str(shopify.get("order_number") or order_id)
    try:
        upstream = await ShiprocketService().find_existing_order(shiprocket_lookup_id)
    except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
        shiprocket_error = str(error)
    upstream_awb = ShiprocketService._upstream_shipment(upstream)[1] if upstream else None
    upstream_status = str(upstream.get("status") or "").strip().casefold() if upstream else ""
    local_awb = shipment.awb if shipment else None
    fulfillment = str(shopify.get("fulfillment_status") or "").casefold()
    protected = bool(local_awb or upstream_awb or upstream_status in {"shipped", "delivered", "in transit", "out for delivery", "pickup scheduled", "ready to ship"} or fulfillment in {"fulfilled", "shipped", "delivered", "in transit", "out for delivery"})
    return {
        "allowed": not protected,
        "shopify": shopify,
        "shiprocket": {"exists": bool(upstream), "order_id": str(upstream.get("id")) if upstream and upstream.get("id") is not None else None, "lookup_id": shiprocket_lookup_id, "status": upstream.get("status") if upstream else None, "awb": upstream_awb, "lookup_error": shiprocket_error},
        "shipment": {"exists": shipment is not None, "provider": shipment.provider if shipment else None, "awb": local_awb, "status": shipment.latest_status if shipment else None},
        "blocked_reason": "Booked/AWB or shipped orders require a separate explicit cancellation workflow." if protected else None,
    }


@router.get("/{order_id}/cancellation/preflight")
async def cancellation_preflight(order_id: str, db: Session = Depends(get_db)) -> dict[str, object]:
    return await _cancellation_preflight(order_id, db)


@router.post("/{order_id}/cancel")
async def cancel_order(order_id: str, payload: CancellationPayload, request: Request, db: Session = Depends(get_db)) -> dict[str, object]:
    preflight = await _cancellation_preflight(order_id, db)
    if not preflight["allowed"]:
        raise HTTPException(status_code=409, detail=preflight["blocked_reason"])
    results: dict[str, object] = {
        "mumchies_os": {"status": "cancelled"},
        "shopify": {"status": "not_applicable"},
        "shiprocket": {"status": "not_applicable"},
        "comment": payload.comment,
    }
    if preflight["shiprocket"].get("lookup_error"):
        results["shiprocket"] = {"status": "failed", "error": preflight["shiprocket"]["lookup_error"], "cancel_on_channel": False}
    if preflight["shopify"].get("cancelled"):
        results["shopify"] = {"status": "already_cancelled"}
    elif payload.cancel_shopify and preflight["shopify"]["exists"]:
        try:
            shopify_result = await ShopifyService().cancel_order(order_id)
            results["shopify"] = shopify_result
        except (ShopifySyncError, httpx.HTTPError) as error:
            results["shopify"] = {"status": "failed", "error": str(error)}
    shiprocket_status = str(preflight["shiprocket"].get("status") or "").strip().casefold()
    if preflight["shiprocket"]["exists"] and shiprocket_status in {"cancelled", "canceled"}:
        results["shiprocket"] = {"status": "Already cancelled", "cancel_on_channel": False}
    elif payload.cancel_shiprocket and preflight["shiprocket"]["exists"]:
        try:
            upstream = await ShiprocketService().find_existing_order(str(preflight["shiprocket"].get("lookup_id") or order_id))
            if upstream:
                cancellation = await ShiprocketService().cancel_unbooked_order(upstream)
                results["shiprocket"] = {"status": "cancelled" if cancellation.get("classification") == "accepted" else "cancellation_requested", "cancel_on_channel": False}
            else:
                results["shiprocket"] = {"status": "failed", "error": "The Shiprocket order disappeared before cancellation.", "cancel_on_channel": False}
        except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
            results["shiprocket"] = {"status": "failed", "error": str(error), "cancel_on_channel": False}
    timestamp = datetime.now(timezone.utc).isoformat()
    actor = current_actor(request)
    operations = OrderOperationsStore.save_cancellation(order_id, results, actor, timestamp)
    return {"results": results, "preflight": preflight, "operations": operations, "timestamp": timestamp, "operator": actor}


def _cleanup_reason(order: ShopifyOrder) -> str | None:
    shipment = order.shipment or {}
    external = order.external_tracking
    provider = str(shipment.get("provider") or (external.provider if external else "")).casefold()
    if provider and provider != "shiprocket" and (shipment.get("awb") or (external.awb if external else None)):
        return "Direct Delhivery shipment" if provider == "delhivery" else f"Booked through {provider}"
    if order.cancelled_at or str(order.shopify_status or "").casefold() == "cancelled":
        return "Cancelled in Shopify"
    if str(order.operational_status or "").casefold() == "cancelled":
        return "Locally marked cancelled"
    if str(order.fulfillment_status or "").casefold() in {"fulfilled", "partial", "partially_fulfilled"}:
        return "Shopify fulfilled"
    return None


@router.get("/shiprocket-cleanup-pending")
async def shiprocket_cleanup_pending(db: Session = Depends(get_db)) -> dict[str, object]:
    orders = await _load_orders(db)
    by_number = {order.order_number: order for order in orders}
    try:
        upstream_orders = await ShiprocketService().list_new_orders()
    except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
        raise HTTPException(status_code=502, detail=str(error)) from error
    if hasattr(db, "scalars"):
        sync_engage_orders(db, {order.order_number: order.order_id for order in orders}, upstream_orders, datetime.now(timezone.utc))
    items = []
    for upstream in upstream_orders:
        number = str(upstream.get("channel_order_id") or "")
        order = by_number.get(number)
        if order is None:
            continue
        reason = _cleanup_reason(order)
        if reason is None:
            continue
        _shipment_id, awb = ShiprocketService._upstream_shipment(upstream)
        if awb:
            continue
        local_shipment = order.shipment or {}
        timeline = OrderOperationsStore.get(order.order_id).get("timeline_events") or []
        last_verification = next((event.get("details") for event in reversed(timeline) if event.get("action") == "shiprocket_cleanup_verified"), None)
        items.append({
            "order_id": order.order_id, "order_number": number,
            "shopify_status": "Cancelled" if order.cancelled_at else order.fulfillment_status or order.shopify_status or "Open",
            "mumchies_provider": local_shipment.get("provider") or (order.external_tracking.provider if order.external_tracking else None),
            "mumchies_status": order.operational_status,
            "shiprocket_order_id": str(upstream.get("id") or ""), "shiprocket_status": upstream.get("status") or "NEW",
            "reason": reason, "shiprocket_awb": None,
            "last_verification": last_verification,
        })
    items.sort(key=lambda value: value["order_number"], reverse=True)
    return {"items": items, "total": len(items)}


def _shiprocket_only_reason(order: ShopifyOrder | None, upstream: dict[str, object]) -> str:
    if order is None:
        return "other"
    if order.cancelled_at or str(order.shopify_status or "").casefold() == "cancelled":
        return "cancelled in Shopify"
    shipment = order.shipment or {}
    provider = str(shipment.get("provider") or (order.external_tracking.provider if order.external_tracking else "")).casefold()
    if provider == "delhivery" and (shipment.get("awb") or (order.external_tracking.awb if order.external_tracking else None)):
        return "fulfilled through Delhivery"
    if provider == "shadowfax" and (shipment.get("awb") or (order.external_tracking.awb if order.external_tracking else None)):
        return "fulfilled through Shadowfax"
    if _has_nested_cancellation_evidence(upstream):
        return "stale Shiprocket state"
    return "other"


def _os_only_reason(order: ShopifyOrder, upstream: dict[str, object] | None) -> str:
    shipment = order.shipment or {}
    provider = str(shipment.get("provider") or (order.external_tracking.provider if order.external_tracking else "")).casefold()
    if provider and provider != "shiprocket" and (shipment.get("awb") or (order.external_tracking.awb if order.external_tracking else None)):
        return "routed directly to another courier"
    if upstream is None:
        return "not yet synced to Shiprocket"
    if str(upstream.get("channel_order_id") or "") != order.order_number:
        return "mapping issue"
    if str(upstream.get("status_code") or "") == "5" or str(upstream.get("status") or "").casefold() in {"cancelled", "canceled"}:
        return "Shiprocket order already cancelled"
    return "other"


def _reconciliation_record(
    *,
    order: ShopifyOrder | None,
    upstream: dict[str, object] | None = None,
    reason: str | None = None,
) -> dict[str, object]:
    upstream = upstream or {}
    if order is not None:
        return {
            "order": order.model_dump(mode="json"),
            "order_id": order.order_id,
            "order_number": order.order_number,
            "created_date": order.created_date,
            "customer_name": order.customer_name,
            "total_amount": float(order.total_amount),
            "payment_type": order.payment_type,
            "risk": "High" if "high" in " ".join(order.tags).casefold() else "Medium" if "medium" in " ".join(order.tags).casefold() else "Low",
            "status": order.operational_status or order.fulfillment_status or order.shopify_status or "Open",
            "reason": reason,
            "shiprocket_order_id": str(upstream.get("id") or "") or None,
            "shiprocket_status": upstream.get("status"),
            "source": "both" if upstream else "os",
        }
    return {
        "order": None,
        "order_id": f"shiprocket:{upstream.get('id')}",
        "order_number": str(upstream.get("channel_order_id") or ""),
        "created_date": upstream.get("created_at") or upstream.get("channel_created_at"),
        "customer_name": upstream.get("customer_name"),
        "total_amount": float(upstream.get("total") or 0),
        "payment_type": "cod" if str(upstream.get("payment_method") or "").casefold() == "cod" else "prepaid",
        "risk": str(upstream.get("rto_risk") or "Low").title(),
        "status": upstream.get("status") or "NEW",
        "reason": reason,
        "shiprocket_order_id": str(upstream.get("id") or "") or None,
        "shiprocket_status": upstream.get("status") or "NEW",
        "source": "shiprocket",
    }


async def _build_reconciliation_summary(db: Session) -> dict[str, object]:
    """Compare active OS work with Shiprocket New without implying the sets must be equal."""
    os_orders = await _load_reconciliation_orders(db)
    operations = [order for order in os_orders if _requires_reconciliation_action(order)]
    service = ShiprocketService()
    try:
        shiprocket_new = await service.list_new_orders(force_refresh=True)
    except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
        raise HTTPException(status_code=502, detail=str(error)) from error
    if hasattr(db, "scalars"):
        sync_engage_orders(db, {order.order_number: order.order_id for order in os_orders}, shiprocket_new, datetime.now(timezone.utc))

    os_by_number: dict[str, list[ShopifyOrder]] = {}
    for order in operations:
        os_by_number.setdefault(order.order_number, []).append(order)
    sr_by_number: dict[str, list[dict[str, object]]] = {}
    for order in shiprocket_new:
        sr_by_number.setdefault(str(order.get("channel_order_id") or ""), []).append(order)

    os_numbers = set(os_by_number)
    sr_numbers = set(sr_by_number)
    both = os_numbers & sr_numbers
    only_os = os_numbers - sr_numbers
    only_sr = sr_numbers - os_numbers
    all_shopify_by_number = {order.order_number: order for order in os_orders}
    lookup_limit = asyncio.Semaphore(5)

    async def classify_os(number: str) -> dict[str, object]:
        order = os_by_number[number][0]
        try:
            async with lookup_limit:
                upstream = await service.find_existing_order(number)
        except (ShiprocketAPIError, ShiprocketConfigurationError):
            upstream = None
            reason = "other"
        else:
            reason = _os_only_reason(order, upstream)
        return {"order_number": number, "reason": reason, "shiprocket_status": (upstream or {}).get("status")}

    only_os_items = await asyncio.gather(*(classify_os(number) for number in sorted(only_os)))
    only_sr_items = [
        {"order_number": number, "reason": _shiprocket_only_reason(all_shopify_by_number.get(number), sr_by_number[number][0]), "shiprocket_status": sr_by_number[number][0].get("status")}
        for number in sorted(only_sr)
    ]
    anomalies = [
        {"order_number": number, "os_records": len(os_by_number.get(number, [])), "shiprocket_records": len(sr_by_number.get(number, []))}
        for number in sorted(os_numbers | sr_numbers)
        if len(os_by_number.get(number, [])) > 1 or len(sr_by_number.get(number, [])) > 1 or not number
    ]
    cleanup_pending = sum(1 for item in only_sr_items if item["reason"] in {"cancelled in Shopify", "fulfilled through Delhivery", "fulfilled through Shadowfax", "stale Shiprocket state"})
    cleanup_reasons = {"cancelled in Shopify", "fulfilled through Delhivery", "fulfilled through Shadowfax", "stale Shiprocket state"}
    datasets = {
        "operations": [_reconciliation_record(order=order, upstream=(sr_by_number.get(order.order_number) or [None])[0]) for order in operations],
        "shiprocket_new": [_reconciliation_record(order=all_shopify_by_number.get(str(upstream.get("channel_order_id") or "")), upstream=upstream) for upstream in shiprocket_new],
        "both": [_reconciliation_record(order=os_by_number[number][0], upstream=sr_by_number[number][0]) for number in sorted(both)],
        "cleanup_pending": [
            _reconciliation_record(order=all_shopify_by_number.get(str(upstream.get("channel_order_id") or "")), upstream=upstream, reason=_shiprocket_only_reason(all_shopify_by_number.get(str(upstream.get("channel_order_id") or "")), upstream))
            for number in sorted(only_sr) for upstream in sr_by_number[number]
            if _shiprocket_only_reason(all_shopify_by_number.get(number), upstream) in cleanup_reasons
        ],
        "missing_in_shiprocket": [
            _reconciliation_record(order=os_by_number[str(item["order_number"])][0], reason=str(item["reason"]))
            for item in only_os_items
        ],
    }
    return {
        "operations_queue": len(operations),
        "fresh_orders": sum(1 for order in operations if _is_fresh_order(order)),
        "previous_pending": sum(1 for order in operations if not _is_fresh_order(order)),
        "shiprocket_new": len(shiprocket_new),
        "present_in_both": len(both),
        "cleanup_pending": cleanup_pending,
        "missing_in_shiprocket": len(only_os),
        "in_both": sorted(both),
        "only_in_os": only_os_items,
        "only_in_shiprocket": only_sr_items,
        "duplicate_mapping_anomalies": anomalies,
        "datasets": datasets,
    }


def _safe_refresh_error(error: Exception) -> str:
    if isinstance(error, HTTPException) and isinstance(error.detail, str):
        return error.detail
    if isinstance(error, (ShopifyConfigurationError, ShiprocketAPIError, ShiprocketConfigurationError)):
        return str(error)
    return "Reconciliation refresh failed. The last successful data is still available."


async def _refresh_reconciliation_snapshot() -> None:
    global _reconciliation_refresh_task
    try:
        with SessionLocal() as db:
            result = await _build_reconciliation_summary(db)
        ReportSnapshotStore.save_success(RECONCILIATION_SNAPSHOT_KEY, result)
    except Exception as error:  # noqa: BLE001 - preserve stale data for any provider/runtime failure
        ReportSnapshotStore.save_error(RECONCILIATION_SNAPSHOT_KEY, _safe_refresh_error(error))
    finally:
        _reconciliation_refresh_task = None


def _start_reconciliation_refresh() -> bool:
    global _reconciliation_refresh_task
    if _reconciliation_refresh_task and not _reconciliation_refresh_task.done():
        return False
    _reconciliation_refresh_task = asyncio.create_task(_refresh_reconciliation_snapshot())
    return True


@router.get("/reconciliation-summary")
async def reconciliation_summary(refresh: bool = False) -> dict[str, object]:
    snapshot = ReportSnapshotStore.get(RECONCILIATION_SNAPSHOT_KEY)
    if refresh or ReportSnapshotStore.is_stale(snapshot, 600):
        _start_reconciliation_refresh()
    if not snapshot or not isinstance(snapshot.get("data"), dict):
        raise HTTPException(status_code=503, detail="Reconciliation is preparing its first snapshot. Try again shortly.")
    return {
        **snapshot["data"],
        "last_refreshed_at": snapshot.get("last_refreshed_at"),
        "refresh_error": snapshot.get("refresh_error"),
        "refreshing": bool(_reconciliation_refresh_task and not _reconciliation_refresh_task.done()),
    }


@router.post("/{order_id}/shiprocket-only-cancel")
async def shiprocket_only_cancel(order_id: str, payload: ShiprocketOnlyCancellationPayload, request: Request) -> dict[str, object]:
    actor = current_actor(request)
    service = ShiprocketService()
    upstream_orders = await service.list_new_orders(force_refresh=True)
    upstream = next((value for value in upstream_orders if str(value.get("id") or "") == payload.shiprocket_order_id and str(value.get("channel_order_id") or "") == payload.order_number), None)
    if upstream is None:
        raise HTTPException(status_code=409, detail="The Shiprocket New order could not be found or no longer matches this order.")
    timestamp = datetime.now(timezone.utc)
    try:
        request_result = await service.request_unbooked_order_cancellation(upstream)
    except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
        request_result = {"http_status": getattr(error, "status_code", None) or 0, "response": {"message": str(error)}, "classification": "rejected"}
    audit = {
        "shiprocket_order_id": payload.shiprocket_order_id,
        "channel_order_id": payload.order_number,
        "request_http_status": request_result["http_status"],
        "request_response": request_result["response"],
        "response_classification": request_result["classification"],
        "operator": actor,
        "timestamp": timestamp.isoformat(),
        "timestamp_ist": timestamp.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M %p IST"),
    }
    OrderOperationsStore.record_timeline_event(order_id, "shiprocket_cleanup_requested", operator=actor, details=audit, timestamp=timestamp.isoformat())
    await asyncio.sleep(0.5)
    return await _verify_shiprocket_only_cancellation(order_id, payload, operator=actor, service=service, request_result=request_result)


def _has_nested_cancellation_evidence(order: dict[str, object] | None) -> bool:
    if not order:
        return False
    products = order.get("products") or []
    activities = order.get("activities") or []
    product_cancelled = any(
        str(product.get("status") or "").casefold() in {"cancelled", "canceled"} or str(product.get("status_code") or "") == "5"
        for product in products if isinstance(product, dict)
    ) if isinstance(products, list) else False
    activity_cancelled = any("cancel" in str(activity).casefold() for activity in activities) if isinstance(activities, list) else False
    return product_cancelled or activity_cancelled


async def _verify_shiprocket_only_cancellation(
    order_id: str,
    payload: ShiprocketOnlyCancellationPayload,
    *,
    operator: str,
    service: ShiprocketService | None = None,
    request_result: dict[str, object] | None = None,
) -> dict[str, object]:
    service = service or ShiprocketService()
    current: dict[str, object] | None = None
    still_in_new: bool | None = None
    verification_error: str | None = None
    try:
        current = await service.find_existing_order(payload.order_number)
        new_orders = await service.list_new_orders(force_refresh=True)
        still_in_new = any(
            str(value.get("id") or "") == payload.shiprocket_order_id
            and str(value.get("channel_order_id") or "") == payload.order_number
            for value in new_orders
        )
    except (ShiprocketAPIError, ShiprocketConfigurationError) as error:
        verification_error = str(error)

    top_status = str((current or {}).get("status") or "") or None
    raw_code = (current or {}).get("status_code")
    try:
        top_code = int(raw_code) if raw_code is not None else None
    except (TypeError, ValueError):
        top_code = None
    top_cancelled = top_code == 5 or str(top_status or "").casefold() in {"cancelled", "canceled"}
    classification = str((request_result or {}).get("classification") or "ambiguous")

    if top_cancelled or still_in_new is False:
        status = "confirmed"
        message = "Shiprocket order cancelled successfully."
    elif str(top_status or "").casefold() == "new" and _has_nested_cancellation_evidence(current):
        status = "inconsistent"
        message = "Shiprocket recorded cancellation activity, but the order remains NEW and can still be shipped."
    elif classification == "rejected":
        status = "rejected"
        message = "Shiprocket rejected the cancellation."
    else:
        status = "unverified"
        message = "Cancellation request was sent, but the final Shiprocket status could not be verified."

    result: dict[str, object] = {
        "status": status,
        "shiprocket_order_id": payload.shiprocket_order_id,
        "channel_order_id": payload.order_number,
        "request_http_status": (request_result or {}).get("http_status"),
        "request_response": (request_result or {}).get("response"),
        "verified_top_level_status": top_status,
        "verified_top_level_status_code": top_code,
        "still_in_new_queue": still_in_new,
        "message": message,
    }
    if verification_error:
        result["verification_error"] = verification_error
    timestamp = datetime.now(timezone.utc)
    audit = {**result, "operator": operator, "timestamp": timestamp.isoformat(), "timestamp_ist": timestamp.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M %p IST")}
    OrderOperationsStore.record_timeline_event(order_id, "shiprocket_cleanup_verified", operator=operator, details=audit, timestamp=timestamp.isoformat())
    return result


@router.post("/{order_id}/shiprocket-only-cancel/verify")
async def verify_shiprocket_only_cancel(order_id: str, payload: ShiprocketOnlyCancellationPayload, request: Request) -> dict[str, object]:
    """Read-only retry: re-check Shiprocket state without sending another cancellation."""
    return await _verify_shiprocket_only_cancellation(order_id, payload, operator=current_actor(request))


@router.post("/{order_id}/address/verify")
async def verify_order_address(order_id: str, payload: VerifyAddressPayload, request: Request) -> dict[str, object]:
    return OrderOperationsStore.verify_address(
        order_id,
        operator=current_actor(request),
        snapshot=payload.address_snapshot,
        verified_at=payload.verified_at or datetime.now().isoformat(timespec="seconds"),
    )
